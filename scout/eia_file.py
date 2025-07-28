#!/usr/bin/env python3
"""Reprocess AEO output files as-received from EIA

Beginning in 2020, EIA modified the structure of some of the NEMS
output files used by Scout. This module restructures those files
to be consistent with the prior format, as expected by the other
modules used to convert the AEO data into the format used by Scout.
"""
import os
import csv
from pathlib import Path
import re
import openpyxl as pyxl
from scout import mseg_techdata as rmt


class EIAFiles(object):
    """AEO data file reprocessing.

    Attributes:
        Original EIA data file names (inputs to this module)
        ----------------------------------------------------
        r_db_in: Residential energy and stock data
        r_mess: Residential equipment and shell component data
        c_tech_in: Commercial technology characteristics data

        Modified EIA data file names (outputs from this module)
        -------------------------------------------------------
        r_db_out: Residential energy and stock data
        r_class: Residential equipment efficiency, lifetime, and adoption
            parameters
        r_meqp: Residential equipment cost, efficiency, and adoption parameters
        c_tech_out: Commercial technology characteristics data

        Other files modified in place
        -----------------------------
        r_lgt: Residential lighting technology data
    """
    def __init__(self, input_dir: Path = None, output_dir: Path = None):
        if not input_dir:
            input_dir = Path(__file__).resolve().parents[1] / 'inputs'
        if not output_dir:
            output_dir = input_dir
        """Define class with required input and output file names."""
        self.input_dir = input_dir
        self.output_dir = output_dir or input_dir

        # ensure the output directory exists
        os.makedirs(self.output_dir, exist_ok=True)

        # inputs
        self.r_db_in = self.input_dir / 'RDM_DBOUT-orig.txt'
        self.r_mess = self.input_dir / 'rsmess.xlsx'
        self.c_tech_in = self.input_dir / 'ktekx.xlsx'
        self.r_lgt_in = self.input_dir / 'rsmlgt.txt'

        # outputs
        self.r_db_out = self.output_dir / 'RDM_DBOUT.txt'
        self.r_class = self.output_dir / 'rsclass.txt'
        self.r_meqp = self.output_dir / 'rsmeqp.txt'
        self.c_tech_out = self.output_dir / 'ktek.csv'
        self.r_lgt_out = self.output_dir / 'rsmlgt.txt'

    def resdbout_fill_household(self):
        """Modify RESDBOUT such that all rows have the same number of columns.

        As delivered by EIA, RESDBOUT has 10 columns, but some rows do
        not have sufficient data or commas inserted into the row for
        all 10 columns. This function updates each row to have entries
        for all 10 columns, including replacing the empty HOUSEHOLDS
        column with a 0.
        """
        # Read the original RESDBOUT data; rename the original
        # RESDBOUT.txt file to RESDBOUT-orig.txt to set it aside if
        # it is not already present so that the modified data can be
        # written to a new RESDBOUT.txt
        try:
            f_dbin = open(self.r_db_in, 'r')
        except FileNotFoundError:
            os.rename(self.r_db_out, self.r_db_in)
            f_dbin = open(self.r_db_in, 'r')

        with open(self.r_db_out, 'w+', encoding='utf-8', newline='') as f_dbout:
            csv_dbin = csv.DictReader(f_dbin)

            # Get field names from the file header row as determined
            # by DictReader
            header = csv_dbin.fieldnames

            # Create corresponding DictWriter object for file outputs
            csv_dbout = csv.DictWriter(f_dbout, fieldnames=header)
            csv_dbout.writeheader()  # Insert header at the top of the new data
            for row in csv_dbin:
                # If households column value is None or empty string,
                # replace with 0 as a string
                if not row['HOUSEHOLDS']:
                    row['HOUSEHOLDS'] = '0'
                # If BULBTYPE column value is None, replace with empty string
                if not row['BULBTYPE']:
                    row['BULBTYPE'] = ''
                # Strip off leading and trailing space characters
                row.update({k: v.strip() for k, v in row.items()})
                csv_dbout.writerow(row)

        f_dbin.close()

    def res_gsl_lt_update(self):
        """Replace 'HAL' with 'INC' for GSL residential lighting.

        Prior to AEO 2020, incandescent general service lighting (GSL)
        bulbs were encoded in the residential data files with the string
        'INC'. This was changed to 'HAL' to indicate the change to more
        efficient halogen incandescent bulbs for GSL bulb types. This
        function reverts the residential data files back to the original
        encoding of 'INC' for incandescent GSL bulbs. This function
        modifies both the lighting technology data file and the
        residential equipment stock and energy use data file
        """
        # Get and revise lighting encoding in residential equipment
        # stock and energy data file
        with open(self.r_db_out, 'r', encoding='utf-8') as f_dbin:
            csv_cont = []
            csv_dbin = csv.DictReader(f_dbin)
            # Get field names from the file header row as determined
            # by DictReader
            header = csv_dbin.fieldnames
            for row in csv_dbin:
                if row.get('EQPCLASS') == 'GSL' and row.get('BULBTYPE') == 'HAL':
                    row['BULBTYPE'] = 'INC'
                csv_cont.append(row)

        # Overwrite residential equipment stock and energy data file
        # with revised lighting strings
        with open(self.r_db_out, 'w', encoding='utf-8', newline='') as f_dbout:
            # Create DictWriter object for file outputs
            csv_dbout = csv.DictWriter(f_dbout, fieldnames=header)
            csv_dbout.writeheader()
            csv_dbout.writerows(csv_cont)

        # update lighting file
        with open(self.r_lgt_in, 'r', encoding='latin1') as f_ltin:
            modified = [
                line.replace('HAL', 'INC')
                if re.search(r'(?<=GSL\s)HAL', line)
                else line
                for line in f_ltin
            ]

        with open(self.r_lgt_out, 'w', encoding='latin1') as f_ltout:
            f_ltout.writelines(modified)

    def generate_rsclass(self):
        """Construct rsclass.txt file from table in combined XLSX file."""
        # derive the sheet name from the filename
        sheet_name = os.path.splitext(
            os.path.basename(self.r_class))[0].upper()
        wb = pyxl.load_workbook(self.r_mess, data_only=True)
        rsclass = wb[sheet_name]

        # Find the position of the "Efficiency Metric" column so that it
        # can be skipped later
        skip = 0  # Placeholder in case 'Efficiency Metric' column is not found
        for cell in rsclass[20]:
            if (cell.value == 'Efficiency Metric'):
                skip = cell.column  # Note that openpyxl is 1-indexed

        with open(self.r_class, 'w+', encoding='utf-8') as f:
            # header row
            cols = [c for c in range(3, 22) if c != skip]
            f.write('\t'.join(str(rsclass.cell(row=20, column=c).value)
                    for c in cols) + '\n')
            # names row
            f.write('\t'.join(rmt.r_nlt_l_names) + '\n')
            # data rows
            for row_num in range(22, 52):
                f.write(
                    '\t'.join(
                        str(rsclass.cell(row=row_num, column=c).value) for c in cols
                    ) + '\n'
                )

    def generate_rsmeqp(self):
        """Construct rsmeqp.txt file from data in combined XLSX file."""
        # derive the sheet name from the filename
        sheet_name = os.path.splitext(os.path.basename(self.r_meqp))[0].upper()
        wb = pyxl.load_workbook(self.r_mess, data_only=True)
        rsmeqp = wb[sheet_name]

        # locate start/end rows
        col_vals = list(*rsmeqp.iter_cols(min_col=2, max_col=2, values_only=True))
        rstart = rend = 0
        for cell in rsmeqp.iter_rows(min_col=2, max_col=2, values_only=False):
            v = cell[0].value
            if v == 'xlI':
                rstart = cell[0].row + 1
            elif isinstance(v, int) and v == list(filter(None, col_vals)).pop():
                rend = cell[0].row + 1

        # Exit function if rend <= rstart
        if rend <= rstart:
            print('rsmeqp table dimensions not obtained successfully.\n')
            print('rsmeqp generation failed.\n')
            return None

        with open(self.r_meqp, 'w+', encoding='utf-8') as f:
            # header
            f.write(
                '\t'.join(
                    str(rsmeqp.cell(row=22, column=c).value) for c in range(3, 31)
                ) + '\n'
            )
            # names
            f.write('\t'.join(rmt.r_nlt_cp_names) + '\n')
            # data
            for row_num in range(rstart, rend):
                f.write(
                    '\t'.join(
                        str(rsmeqp.cell(row=row_num, column=c).value) for c in range(3, 31)
                    ) + '\n'
                )

    def convert_ktekx(self):
        """Convert commercial technology characteristics data to CSV format."""
        if not os.path.isfile(self.c_tech_out):
            sheet_name = os.path.splitext(os.path.basename(self.c_tech_out))[0]
            ktekx = pyxl.load_workbook(
                self.c_tech_in, data_only=True)[sheet_name]
            with open(self.c_tech_out, 'w+', newline='') as f:
                writer = csv.writer(f)
                for row_n, row in enumerate(ktekx.rows):
                    # Pull row values to write
                    row_vals = [cell.value for cell in row]
                    # Ensure that the final row includes values to write (not empty line)
                    if (row_n + 1) != ktekx.max_row or (((row_n + 1) == ktekx.max_row) and not all(
                            [x is None for x in row_vals])):
                        writer.writerow(row_vals)
        else:
            print(f"{self.c_tech_out} is already present and will not be modified.")


def main():
    # read from AEO raw data, write processed files into 'inputs'
    f = EIAFiles(input_dir=Path('inputs'), output_dir=Path('inputs'))

    f.resdbout_fill_household()
    f.res_gsl_lt_update()
    f.generate_rsclass()
    f.generate_rsmeqp()
    f.convert_ktekx()


if __name__ == "__main__":
    main()
