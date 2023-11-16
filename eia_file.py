#!/usr/bin/env python3
"""Reprocess AEO output files as-received from EIA

Beginning in 2020, EIA modified the structure of some of the NEMS
output files used by Scout. This module restructures those files
to be consistent with the prior format, as expected by the other
modules used to convert the AEO data into the format used by Scout.
"""
import os
import csv
import re
import openpyxl as pyxl
import mseg_techdata as rmt


class EIAFiles(object):
    """AEO data file reprocessing.

    Attributes:
        Original EIA data file names (inputs to this module)
        ----------------------------------------------------
        r_db_in: Residential energy and stock data
        r_mess: Residential equipment and shell component data
        c_tech_in: Commercial technology characteristics data

        Modified EIA data file names (outputs from this module)
        -------------------------------------------------------
        r_db_out: Residential energy and stock data
        r_class: Residential equipment efficiency, lifetime, and adoption
            parameters
        r_meqp: Residential equipment cost, efficiency, and adoption parameters
        c_tech_out: Commercial technology characteristics data

        Other files modified in place
        -----------------------------
        r_lgt: Residential lighting technology data
    """
    def __init__(self):
        """Define class with required input and output file names."""
        self.r_db_in = 'RESDBOUT-orig.txt'
        self.r_mess = 'rsmess.xlsx'
        self.r_db_out = 'RESDBOUT.txt'
        self.r_class = 'rsclass.txt'
        self.r_meqp = 'rsmeqp.txt'
        self.c_tech_in = 'ktekx.xlsx'
        self.c_tech_out = 'ktek.csv'
        self.r_lgt = 'rsmlgt.txt'

    def resdbout_fill_household(self):
        """Modify RESDBOUT such that all rows have the same number of columns.

        As delivered by EIA, RESDBOUT has 10 columns, but some rows do
        not have sufficient data or commas inserted into the row for
        all 10 columns. This function updates each row to have entries
        for all 10 columns, including replacing the empty HOUSEHOLDS
        column with a 0.
        """
        # Read the original RESDBOUT data; rename the original
        # RESDBOUT.txt file to RESDBOUT-orig.txt to set it aside if
        # it is not already present so that the modified data can be
        # written to a new RESDBOUT.txt
        try:
            f_dbin = open(self.r_db_in, 'r')
        except FileNotFoundError:
            os.rename(self.r_db_out, self.r_db_in)
            f_dbin = open(self.r_db_in, 'r')

        with open(self.r_db_out, 'w+', encoding='utf-8', newline='') as f_dbout:
            csv_dbin = csv.DictReader(f_dbin)

            # Get field names from the file header row as determined
            # by DictReader
            header = csv_dbin.fieldnames

            # Create corresponding DictWriter object for file outputs
            csv_dbout = csv.DictWriter(f_dbout, fieldnames=header)
            csv_dbout.writeheader()  # Insert header at the top of the new data
            for row in csv_dbin:
                # If households column value is None or empty string,
                # replace with 0 as a string
                if not row['HOUSEHOLDS']:
                    row['HOUSEHOLDS'] = '0'
                # If BULBTYPE column value is None, replace with empty string
                if not row['BULBTYPE']:
                    row['BULBTYPE'] = ''
                # Strip off leading and trailing space characters
                row.update({k: v.strip() for k, v in row.items()})

                csv_dbout.writerow(row)

        f_dbin.close()

    def res_gsl_lt_update(self):
        """Replace 'HAL' with 'INC' for GSL residential lighting.

        Prior to AEO 2020, incandescent general service lighting (GSL)
        bulbs were encoded in the residential data files with the string
        'INC'. This was changed to 'HAL' to indicate the change to more
        efficient halogen incandescent bulbs for GSL bulb types. This
        function reverts the residential data files back to the original
        encoding of 'INC' for incandescent GSL bulbs. This function
        modifies both the lighting technology data file and the
        residential equipment stock and energy use data file
        """
        # Get and revise lighting encoding in residential equipment
        # stock and energy data file
        with open(self.r_db_out, 'r') as f_dbin:
            csv_cont = []
            csv_dbin = csv.DictReader(f_dbin)
            # Get field names from the file header row as determined
            # by DictReader
            header = csv_dbin.fieldnames
            for row in csv_dbin:
                if row['EQPCLASS'] == 'GSL':
                    if row['BULBTYPE'] == 'HAL':
                        row['BULBTYPE'] = 'INC'
                csv_cont.append(row)

        # Overwrite residential equipment stock and energy data file
        # with revised lighting strings
        with open(self.r_db_out, 'w', encoding='utf-8', newline='') as f_dbout:
            # Create DictWriter object for file outputs
            csv_dbout = csv.DictWriter(f_dbout, fieldnames=header)
            csv_dbout.writeheader()
            csv_dbout.writerows(csv_cont)

        # Get and revise lighting encoding in residential lighting
        # technology characteristics file
        with open(self.r_lgt, 'r', encoding='latin1') as f_ltin:
            f_cont = []
            for line in f_ltin:
                if re.search(r'(?<=GSL\s)HAL', line):
                    line = line.replace('HAL', 'INC')
                f_cont.append(line)

        # Overwrite residential lighting characteristics data file
        # with revised lighting string
        with open(self.r_lgt, 'w', encoding='latin1') as f_ltout:
            f_ltout.writelines(f_cont)

    def generate_rsclass(self):
        """Construct rsclass.txt file from table in combined XLSX file."""

        # Load applicable worksheet from the spreadsheet file
        sheet_name = self.r_class.split('.')[0].upper()
        rsclass = pyxl.load_workbook(self.r_mess, data_only=True)[sheet_name]

        # Find the position of the "Efficiency Metric" column so that it
        # can be skipped later
        skip = 0  # Placeholder in case 'Efficiency Metric' column is not found
        for cell in rsclass[19]:
            if (cell.value == 'Efficiency Metric'):
                skip = cell.column  # Note that openpyxl is 1-indexed

        with open(self.r_class, 'w+', encoding='utf-8') as f:
            # Construct and set header row
            f.write('\t'.join(str(rsclass.cell(row=19, column=col_num).value)
                              for col_num in [x for x in range(3, 22) if x != skip]) + '\n')
            # Construct and set subheading
            f.write('\t'.join(str(name) for name in rmt.r_nlt_l_names) + '\n')
            # Populate body rows
            for row_num in range(21, 51):
                f.write('\t'.join(str(rsclass.cell(row=row_num, column=col_num).value)
                                  for col_num in [x for x in range(3, 22) if x != skip]) + '\n')
            f.close()

    def generate_rsmeqp(self):
        """Construct rsmeqp.txt file from data in combined XLSX file."""

        # Load applicable worksheet from the spreadsheet file
        sheet_name = self.r_meqp.split('.')[0].upper()
        rsmeqp = pyxl.load_workbook(self.r_mess, data_only=True)[sheet_name]

        # Set starting and ending table row variables with placeholder values
        rstart = rend = 0
        # Get entries in first column as a flat list
        col_1 = list(*rsmeqp.iter_cols(max_col=1, values_only=True))
        # Find the starting and ending rows of the table in the worksheet
        # based on the entries in the first column
        for col in rsmeqp.iter_cols(max_col=1):
            for r in col:
                if r.value == 'xlI':
                    rstart = r.row + 1  # Add 1 to skip model variable names row
                # Second term gets the last non-None value
                elif isinstance(r.value, int) and r.value == list(filter(None, col_1)).pop():
                    rend = r.row + 1

        # Exit function if rend <= rstart
        if rend <= rstart:
            print('rsmeqp table dimensions not obtained successfully.\n')
            print('rsmeqp generation failed.\n')
            return None

        with open(self.r_meqp, 'w+') as f:
            # Construct and set header row
            f.write('\t'.join(str(rsmeqp.cell(row=22, column=col_num).value)
                              for col_num in range(2, 30)) + '\n')
            # Construct and set subheading
            f.write('\t'.join(str(name) for name in rmt.r_nlt_cp_names) + '\n')
            # Populate body rows
            for row_num in range(rstart, rend):
                f.write('\t'.join(str(rsmeqp.cell(row=row_num, column=col_num).value)
                                  for col_num in range(2, 30)) + '\n')
        f.close()

    def convert_ktekx(self):
        """Convert commercial technology characteristics data to CSV format."""
        if not os.path.isfile(self.c_tech_out):
            # Load applicable worksheet from the spreadsheet file
            sheet_name = self.c_tech_out.split('.')[0]
            ktekx = pyxl.load_workbook(self.c_tech_in, data_only=True)[sheet_name]

            # Create target CSV file
            ktek = csv.writer(open(self.c_tech_out, 'w+'))

            # Stream content to CSV file
            for r in ktekx.rows:
                row = [a.value for a in r]
                ktek.writerow(row)
        else:
            print('ktek.csv is already present and will not be modified.')


def main():
    f = EIAFiles()

    f.resdbout_fill_household()
    f.res_gsl_lt_update()
    f.generate_rsclass()
    f.generate_rsmeqp()
    f.convert_ktekx()


if __name__ == "__main__":
    main()
